"""Excel (.xlsx/.xls) loader using openpyxl. Each sheet becomes a page_map entry."""
from __future__ import annotations
import uuid
from pathlib import Path

from raglib.loaders.base import BaseLoader
from raglib.models.document import Document


class ExcelLoader(BaseLoader):
    """Loads Excel files. Each worksheet is converted to readable text and becomes a page."""

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]

    def load(self, path: str) -> list[Document]:
        try:
            import openpyxl  # type: ignore
        except ImportError:
            raise ImportError(
                "openpyxl is required to load Excel files. "
                "Install it with: pip install openpyxl"
            )

        p = Path(path)
        doc_id = p.stem + "_" + uuid.uuid4().hex[:8]

        # openpyxl data_only=True reads cached formula values
        wb = openpyxl.load_workbook(str(p), data_only=True)

        text_parts: list[str] = []
        page_map: list[dict] = []
        char_offset = 0

        for sheet_num, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            sheet_text = self._sheet_to_text(ws, sheet_name)

            separator = "\n\n" if text_parts else ""
            full_part = separator + sheet_text

            start_char = char_offset + len(separator)
            end_char = char_offset + len(full_part)

            page_map.append({
                "page_num": sheet_num,
                "start_char": start_char,
                "end_char": end_char,
                "heading": sheet_name,
            })

            text_parts.append(full_part)
            char_offset = end_char

        full_text = "".join(text_parts)

        return [
            Document(
                doc_id=doc_id,
                text=full_text,
                source_path=str(p.resolve()),
                source_name=p.name,
                loader="ExcelLoader",
                file_extension=p.suffix.lower(),
                file_size_bytes=p.stat().st_size,
                page_count=len(page_map),
                page_map=page_map,
            )
        ]

    def _sheet_to_text(self, ws, sheet_name: str) -> str:
        """Convert a worksheet to a readable text block."""
        lines = [f"Sheet: {sheet_name}"]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            lines.append("(empty)")
            return "\n".join(lines)

        # First row is treated as the header
        header = rows[0]
        header_str = "\t".join("" if v is None else str(v) for v in header)
        lines.append(header_str)

        for row in rows[1:]:
            row_str = "\t".join("" if v is None else str(v) for v in row)
            lines.append(row_str)

        return "\n".join(lines)
